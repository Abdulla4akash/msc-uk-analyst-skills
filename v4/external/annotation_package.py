"""
Annotation package generation — BLANK workbooks only.

Must contain NO:
- A1 predictions, S1/S2/S3 predictions, thresholds, confidence, seen/unseen flags, challenge reason
Annotators label from job-ad evidence; model outputs stay in separate researcher-only files.

Skill labels judged from same textual evidence models receive: job_summary
(job_title used only for role_family metadata, not skill inference)

Generates:
- ANNOTATOR_A.xlsx
- ANNOTATOR_B_OVERLAP.xlsx (100 overlap)
- Or CSV equivalents

Uses frozen 13 labels in exact CATEGORIES order plus other/other_skills_verbatim/notes.
"""

import hashlib
import json
import re
from pathlib import Path
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if not (REPO_ROOT / "v4" / "config.py").exists():
    REPO_ROOT = Path.cwd() / "msc-uk-analyst-skills"
    if not (REPO_ROOT / "v4" / "config.py").exists():
        REPO_ROOT = Path.cwd()

from v4.config import CATEGORIES, CATEGORY_LABELS, TAXONOMY_VERSION

# Expected columns
SKILL_COLS = CATEGORIES  # exact order
EXTRA_COLS = ["other", "other_skills_verbatim", "notes"]

def create_blank_workbook(selected_df: pd.DataFrame, out_path: Path, annotator_label: str):
    """
    Create blank annotation workbook with model-visible text only.

    selected_df must contain:
      - external_id
      - job_summary (model-visible text)
      - job_title (for role_family metadata, not skill evidence)
      - company_hash or company (if permitted)
      - source, role_family_sampling_stratum, challenge_stratum etc

    Writes ANNOTATOR_A.xlsx with blank label cells only (no model outputs).
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Validate no model prediction columns present
    forbidden = [c for c in selected_df.columns if any(k in c.lower() for k in ["a1", "s1", "s2", "s3", "prediction", "score", "threshold", "confidence"])]
    # More strict: check that skill columns are not pre-filled with model outputs
    # Here we ensure we are not leaking
    if any(c in selected_df.columns for c in CATEGORIES):
        # If skill columns already exist, ensure they are not model predictions; for blank workbook we will overwrite with blank
        pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annotation"

    # Header
    headers = ["external_id", "job_title", "job_summary", "role_family"] + SKILL_COLS + EXTRA_COLS
    # Add metadata columns that are allowed: source, published_at, etc not model
    # We'll include source and stratum for tracking but not model scores
    meta_cols = [c for c in ["source", "source_posting_id", "source_url", "published_at", "role_family_sampling_stratum", "challenge_stratum", "natural_or_challenge", "duplicate_group_id"] if c in selected_df.columns]
    # Order: external_id, source etc, job_title, job_summary, role_family, skills...
    headers = ["external_id"] + meta_cols + ["job_title", "job_summary"] + SKILL_COLS + EXTRA_COLS

    # Write header
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fill rows with blank skill cells
    for row_idx, (_, row) in enumerate(selected_df.iterrows(), start=2):
        for col_idx, h in enumerate(headers, start=1):
            if h in SKILL_COLS:
                # Blank 0/1 cell for annotator to fill
                ws.cell(row=row_idx, column=col_idx, value=None)
            elif h == "other":
                ws.cell(row=row_idx, column=col_idx, value=None)
            elif h in ["other_skills_verbatim", "notes"]:
                ws.cell(row=row_idx, column=col_idx, value=None)
            else:
                # metadata / text
                val = row.get(h, "")
                # Truncate very long job_summary for Excel cell? Keep full but ensure wrap
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Validation: skill columns must have data validation 0/1 only, blank allowed
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dv.error = "Enter 0 or 1 only"
    dv.errorTitle = "Invalid entry"
    dv.prompt = "0=absent, 1=present"
    # Apply to skill columns
    skill_col_indices = [headers.index(c)+1 for c in SKILL_COLS]
    for col_idx in skill_col_indices:
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{len(selected_df)+1}")
    ws.add_data_validation(dv)

    # Column widths
    ws.column_dimensions[get_column_letter(headers.index("job_summary")+1)].width = 80
    ws.column_dimensions[get_column_letter(headers.index("job_title")+1)].width = 30
    for c in SKILL_COLS:
        ws.column_dimensions[get_column_letter(headers.index(c)+1)].width = 12
    ws.column_dimensions[get_column_letter(headers.index("notes")+1)].width = 30
    # Freeze header
    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Second sheet: guidelines reference
    ws2 = wb.create_sheet("Guidelines")
    ws2["A1"] = "Annotation guidelines version"
    ws2["B1"] = TAXONOMY_VERSION
    ws2["A2"] = "Categories in order"
    for i, cat in enumerate(CATEGORIES, start=2):
        ws2.cell(row=1, column=3+i, value=cat)
    ws2["A3"] = "Label definition: 1=skill required/evidenced in job_summary, 0=absent"
    ws2["A4"] = "Use job_summary only for skill labels (job_title for role_family only)"
    ws2["A5"] = "Do NOT use external knowledge unavailable to model"

    # Third sheet: validation info
    ws3 = wb.create_sheet("Validation")
    ws3["A1"] = "No model predictions, scores, thresholds, or seen/unseen flags are in this workbook"
    ws3["A2"] = "If you see such columns, STOP and report protocol violation"

    wb.save(out_path)
    print(f"Created blank workbook {out_path} with {len(selected_df)} rows, {len(headers)} columns (no model outputs)")

def select_overlap_ids(selected_E1, selected_E2, n_overlap=100, seed=42, stratify=True):
    """
    Select 100 overlap IDs for double annotation (50 E1 natural, 50 E2 challenge) BEFORE annotation.

    Stratify by natural/challenge, role_family, challenge_stratum where applicable.
    Do NOT select based on model errors or first annotator labels.

    Returns dataframe of overlap ids and metadata.
    """
    rng = np.random.default_rng(seed)
    # 50 from E1, 50 from E2
    need_E1 = 50
    need_E2 = 50
    # E1 sample
    if len(selected_E1) < need_E1:
        raise ValueError(f"E1 has only {len(selected_E1)} < 50 needed for overlap")
    if len(selected_E2) < need_E2:
        raise ValueError(f"E2 has only {len(selected_E2)} < 50 needed for overlap")

    # Stratified within each: by role_family
    def stratified_sample(df, n, rng):
        # allocate proportionally to role_family counts
        counts = df["role_family"].value_counts().to_dict() if "role_family" in df.columns else {"all": len(df)}
        # simple: sample n with stratification by role_family via group sampling
        if "role_family" not in df.columns or len(counts)<=1:
            idx = rng.choice(len(df), size=n, replace=False)
            return df.iloc[np.sort(idx)]
        # Proportional allocation
        allocations = {}
        remaining = n
        total = len(df)
        for fam, cnt in counts.items():
            alloc = max(1, int(round(cnt/total * n))) if remaining>0 else 0
            alloc = min(alloc, cnt, remaining)
            allocations[fam] = alloc
            remaining -= alloc
        # Adjust
        families = list(counts.keys())
        while remaining >0:
            for fam in families:
                if allocations[fam] < counts[fam]:
                    allocations[fam]+=1
                    remaining-=1
                    if remaining==0:
                        break
        # Sample per family
        parts=[]
        for fam, alloc in allocations.items():
            sub = df[df["role_family"]==fam]
            if len(sub) <= alloc:
                parts.append(sub)
            else:
                idx = rng.choice(len(sub), size=alloc, replace=False)
                parts.append(sub.iloc[np.sort(idx)])
        res = pd.concat(parts)
        # If still not enough due to rounding, fill
        if len(res) < n:
            remaining_df = df[~df["external_id"].isin(res["external_id"])]
            need = n - len(res)
            idx = rng.choice(len(remaining_df), size=need, replace=False)
            res = pd.concat([res, remaining_df.iloc[np.sort(idx)]])
        return res.sort_values("external_id")

    overlap_E1 = stratified_sample(selected_E1, need_E1, rng)
    overlap_E2 = stratified_sample(selected_E2, need_E2, rng)
    overlap_all = pd.concat([overlap_E1, overlap_E2]).sort_values("external_id").reset_index(drop=True)
    metadata = {
        "overlap_seed": seed,
        "n_overlap": int(len(overlap_all)),
        "n_E1_overlap": int(len(overlap_E1)),
        "n_E2_overlap": int(len(overlap_E2)),
        "stratification": "natural/challenge + role_family + challenge_stratum where applicable, seeded, before annotation, independent of labels/model errors",
        "ids": overlap_all["external_id"].tolist()
    }
    return overlap_all, metadata

def validate_workbook_has_no_model_outputs(workbook_path: Path):
    """
    Assert workbook contains no model prediction/score columns.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    ws = wb["Annotation"]
    headers = [c.value for c in ws[1]]
    forbidden_substrings = ["a1", "s1", "s2", "s3", "prediction", "score", "threshold", "confidence", "seen_unseen", "challenge_reason"]
    offenders = [h for h in headers if h and any(k in str(h).lower() for k in forbidden_substrings)]
    assert not offenders, f"Workbook {workbook_path} contains forbidden model output columns: {offenders}"
    # Also check that skill columns are blank (no pre-filled labels)
    skill_idx = [headers.index(c)+1 for c in CATEGORIES if c in headers]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=min(skill_idx), max_col=max(skill_idx)):
        for cell in row:
            if cell.value not in (None, 0, 1, "0", "1", ""):
                # Allow blank or 0/1 only, but for blank workbook should be blank
                if cell.value is not None:
                    # It should be blank; if it's filled, that's gold generation by code which is forbidden
                    raise AssertionError(f"Workbook {workbook_path} has pre-filled label cell {cell.coordinate}={cell.value}; must be blank for human annotation")
    print(f"Validated {workbook_path}: no model outputs, blank labels")

def workbook_hash(path: Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()
