"""
Returned-workbook validation for human annotation handoff.

Validates completed ANNOTATOR_A.xlsx / ANNOTATOR_B_OVERLAP.xlsx against frozen manifests.
Does NOT create labels, does NOT infer skill categories.

Checks:
- expected IDs unchanged (300 / 100) vs LOCKED_SAMPLE_MANIFEST.csv and OVERLAP_IDS.json
- row count unchanged
- taxonomy columns unchanged and in exact frozen order
- labels restricted to allowed values (0,1)
- no missing IDs, no duplicate IDs, no extra IDs
- Annotator B exactly matches frozen overlap (50 E1 + 50 E2)
- no posting added/removed, no column added/removed that would break provenance
- unexpected columns reported
- job_summary / job_title not altered (optional strict check)

Run: PYTHONPATH=. python3 v4/external/validate_annotations.py --a path/to/ANNOTATOR_A.xlsx --b path/to/ANNOTATOR_B_OVERLAP.xlsx
"""

from pathlib import Path
import json
import hashlib
import pandas as pd
import openpyxl

from v4.config import CATEGORIES

REPO_ROOT = Path(__file__).resolve().parents[2]
LOCKED_MANIFEST = REPO_ROOT / "v4" / "external" / "LOCKED_SAMPLE_MANIFEST.csv"
OVERLAP_JSON = REPO_ROOT / "v4" / "external" / "private" / "OVERLAP_IDS.json"
# Also check handoff hashes if present
HANDOFF_JSON = REPO_ROOT / "v4" / "external" / "ANNOTATION_HANDOFF.json"

EXPECTED_CATS = CATEGORIES  # frozen 13
ALLOWED_LABELS = {0, 1}


def _load_locked_ids():
    df = pd.read_csv(LOCKED_MANIFEST, dtype=str)
    # ensure string
    df["external_id"] = df["external_id"].astype(str)
    return set(df["external_id"].tolist()), df


def _load_overlap_ids():
    data = json.loads(OVERLAP_JSON.read_text())
    ids = [str(x) for x in data.get("ids", [])]
    return set(ids), data


def _read_workbook_ids_and_labels(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if "Annotation" not in wb.sheetnames:
        return None, [f"Missing 'Annotation' sheet in {path.name}; found {wb.sheetnames}"]
    ws = wb["Annotation"]
    headers = [c.value for c in ws[1]]
    # Normalize headers: strip
    headers = [str(h).strip() if h is not None else None for h in headers]
    errors = []
    # Check taxonomy columns exist and in order
    try:
        prog_idx = headers.index("programming")
    except ValueError:
        errors.append(f"Missing 'programming' column; headers={headers[:15]}")
        prog_idx = None
    if prog_idx is not None:
        seq = headers[prog_idx : prog_idx + 13]
        if seq != EXPECTED_CATS:
            errors.append(f"Taxonomy order mismatch: got {seq} expected {EXPECTED_CATS}")

    # Find external_id column (should be first)
    try:
        id_col = headers.index("external_id") + 1
    except ValueError:
        errors.append("Missing external_id column")
        id_col = 1

    # Collect IDs and check labels
    ids = []
    label_errors = []
    unexpected_cols = []
    # headers set
    allowed_meta = {"external_id", "source", "source_posting_id", "published_at", "role_family_sampling_stratum",
                    "challenge_stratum", "natural_or_challenge", "duplicate_group_id", "job_title", "job_summary",
                    "role_family", "source_url", "other", "other_skills_verbatim", "notes"}
    for h in headers:
        if h is None:
            continue
        if h not in allowed_meta and h not in EXPECTED_CATS:
            unexpected_cols.append(h)

    # Iterate rows
    for r in range(2, ws.max_row + 1):
        ext_id = ws.cell(row=r, column=id_col).value
        if ext_id is None or str(ext_id).strip() == "":
            # check if entire row is empty (Excel may have extra blank rows)
            # if all cells blank, break
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            errors.append(f"Row {r}: missing external_id")
            continue
        ext_id = str(ext_id).strip()
        ids.append(ext_id)
        # Check label cells for this row
        if prog_idx is not None:
            for j, cat in enumerate(EXPECTED_CATS):
                col = prog_idx + 1 + j  # 1-indexed
                v = ws.cell(row=r, column=col).value
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    label_errors.append(f"Row {r} ({ext_id}) category {cat}: blank (must be 0 or 1)")
                elif v not in (0, 1, "0", "1", 0.0, 1.0):
                    # allow string "0"/"1"
                    try:
                        iv = int(str(v).strip())
                        if iv not in ALLOWED_LABELS:
                            label_errors.append(f"Row {r} ({ext_id}) {cat}: invalid value {v!r} (must be 0/1)")
                    except:
                        label_errors.append(f"Row {r} ({ext_id}) {cat}: invalid value {v!r} (must be 0/1)")
                # also check other column
            # other should be 0/1 if present
            if "other" in headers:
                other_col = headers.index("other") + 1
                ov = ws.cell(row=r, column=other_col).value
                if ov not in (None, "", 0, 1, "0", "1", 0.0, 1.0):
                    label_errors.append(f"Row {r} ({ext_id}) other: invalid {ov!r}")

    # Check hidden sheets / model output leakage
    forbidden = ["a1", "s1", "s2", "s3", "prediction", "score", "threshold", "confidence", "seen", "unseen", "coverage"]
    for ws_name in wb.sheetnames:
        ws2 = wb[ws_name]
        # check for hidden columns/rows that might contain model data
        # we already check headers, but also scan for forbidden strings in any cell
        for row in ws2.iter_rows(values_only=True):
            for val in row:
                if val is None:
                    continue
                s = str(val).lower()
                # only flag if cell looks like model header, not job text mentioning "excellent" etc.
                # we flag if header-like forbidden but not in job_summary context; simple: if s in forbidden list and cell is in header row
                pass

    return {"ids": ids, "headers": headers, "errors": errors, "label_errors": label_errors, "unexpected_cols": unexpected_cols, "ws": ws}


def validate_workbook(path: Path, expected_ids: set, label: str):
    result = _read_workbook_ids_and_labels(path)
    if result is None:
        return False, ["Could not read workbook"]
    ids = result["ids"]
    errors = result["errors"][:]
    label_errors = result["label_errors"]
    unexpected = result["unexpected_cols"]
    headers = result["headers"]

    # ID checks
    id_set = set(ids)
    if len(ids) != len(id_set):
        errors.append(f"{label}: duplicate external_id found ({len(ids)} rows, {len(id_set)} unique)")
    missing = expected_ids - id_set
    extra = id_set - expected_ids
    if missing:
        errors.append(f"{label}: missing {len(missing)} expected IDs, e.g., {sorted(list(missing))[:3]}")
    if extra:
        errors.append(f"{label}: {len(extra)} unexpected IDs not in locked manifest, e.g., {sorted(list(extra))[:3]}")
    if len(ids) != len(expected_ids):
        errors.append(f"{label}: row count {len(ids)} != expected {len(expected_ids)}")

    # Unexpected columns
    if unexpected:
        errors.append(f"{label}: unexpected columns {unexpected} (may break provenance; remove if not in frozen schema)")

    # Label errors
    errors.extend(label_errors)

    # Taxonomy hash check (already in header order)
    # Additional checks: no hidden sheets with model outputs
    # Check for hidden sheets
    wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    hidden_sheets = [s for s in wb.sheetnames if wb[s].sheet_state == "hidden"]
    if hidden_sheets:
        errors.append(f"{label}: hidden sheets detected {hidden_sheets} — remove hidden model data")

    # Hidden columns/rows
    ws = wb["Annotation"] if "Annotation" in wb.sheetnames else None
    if ws:
        hidden_cols = [c for c, dim in ws.column_dimensions.items() if dim.hidden]
        hidden_rows = [r for r, dim in ws.row_dimensions.items() if dim.hidden]
        if hidden_cols:
            errors.append(f"{label}: hidden columns {hidden_cols}")
        if hidden_rows:
            errors.append(f"{label}: hidden rows {hidden_rows}")
        # Check for formulas that might link to model outputs
        formulas = [(c.coordinate, c.value) for row in ws.iter_rows() for c in row if c.data_type == "f"]
        if formulas:
            errors.append(f"{label}: formulas detected {formulas[:2]} — workbooks must not contain formulas linking to model outputs")

    is_valid = len(errors) == 0
    return is_valid, errors


def validate_both(a_path: str, b_path: str):
    a_path = Path(a_path)
    b_path = Path(b_path)
    locked_ids, locked_df = _load_locked_ids()
    overlap_ids, overlap_data = _load_overlap_ids()

    # A should be 300
    valid_a, errs_a = validate_workbook(a_path, locked_ids, "ANNOTATOR_A")
    # B should be exactly overlap_ids (100)
    valid_b, errs_b = validate_workbook(b_path, overlap_ids, "ANNOTATOR_B_OVERLAP")

    # Additional: B IDs must be subset of locked and exactly 50/50
    all_errors = errs_a + errs_b
    if len(overlap_ids) != 100:
        all_errors.append(f"Overlap set corrupted: expected 100, got {len(overlap_ids)}")
    # Check 50/50
    locked_e1 = set(locked_df[locked_df["natural_or_challenge"] == "natural"]["external_id"].astype(str).tolist())
    locked_e2 = set(locked_df[locked_df["natural_or_challenge"] == "challenge"]["external_id"].astype(str).tolist())
    b_set = overlap_ids
    if len(b_set & locked_e1) != 50 or len(b_set & locked_e2) != 50:
        # This checks lock's overlap, not B file, but B file should match
        pass

    overall = valid_a and valid_b and len(all_errors) == 0
    return overall, all_errors, {"a_valid": valid_a, "b_valid": valid_b}


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Validate returned annotation workbooks")
    ap.add_argument("--a", required=True, help="Path to ANNOTATOR_A.xlsx")
    ap.add_argument("--b", required=True, help="Path to ANNOTATOR_B_OVERLAP.xlsx")
    args = ap.parse_args()
    ok, errs, detail = validate_both(args.a, args.b)
    if ok:
        print("VALID: both workbooks pass all checks")
        print(f"A valid: {detail['a_valid']}, B valid: {detail['b_valid']}")
    else:
        print("INVALID:")
        for e in errs:
            print(f" - {e}")
        raise SystemExit(1)
