"""
Tests for human annotation handoff — verifies locked sample integrity, workbook blinding, and validator.
All tests use synthetic data only; no real human labels are fabricated.
"""

import hashlib
import json
from pathlib import Path
import pandas as pd
import openpyxl
import pytest

from v4.config import CATEGORIES
from v4.tests._paths import REPO_ROOT

LOCKED = REPO_ROOT / "v4" / "external" / "LOCKED_SAMPLE_MANIFEST.csv"
PRIVATE = REPO_ROOT / "v4" / "external" / "private" / "LOCKED_SAMPLE_MANIFEST_PRIVATE.csv"
LOCK_JSON = REPO_ROOT / "v4" / "external" / "EXTERNAL_SAMPLE_LOCK.json"
HANDOFF = REPO_ROOT / "v4" / "external" / "ANNOTATION_HANDOFF.json"
OVERLAP_JSON = REPO_ROOT / "v4" / "external" / "private" / "OVERLAP_IDS.json"
WORKBOOK_A = REPO_ROOT / "v4" / "external" / "private" / "annotation" / "ANNOTATOR_A.xlsx"
WORKBOOK_B = REPO_ROOT / "v4" / "external" / "private" / "annotation" / "ANNOTATOR_B_OVERLAP.xlsx"
GUIDE = REPO_ROOT / "v4" / "external" / "HUMAN_ANNOTATION_GUIDE.md"
CHECKLIST = REPO_ROOT / "v4" / "external" / "ANNOTATOR_CHECKLIST.md"


def test_locked_manifest_hash_unchanged():
    # Expected hash from lock file (previous hand-off)
    expected = "884c521383b1581df75241843b10e98d4c1954539820543592b8e90734034bf3"
    h = hashlib.sha256(LOCKED.read_bytes()).hexdigest()
    assert h == expected, f"Manifest hash changed {h} != {expected}"
    # Also compare to lock json
    lock = json.loads(LOCK_JSON.read_text())
    assert lock["locked_sample_manifest_sha256"] == expected
    # Also handoff
    if HANDOFF.exists():
        hand = json.loads(HANDOFF.read_text())
        assert hand["locked_sample_manifest_sha256"] == expected


def test_public_private_manifest_alignment():
    df = pd.read_csv(LOCKED, dtype=str)
    dfp = pd.read_csv(PRIVATE, dtype=str)
    assert len(df) == 300
    assert len(dfp) == 300
    assert set(df["external_id"]) == set(dfp["external_id"])
    # shared safe fields
    shared = ["source", "source_posting_id", "published_at", "acquired_at",
              "role_family_sampling_stratum", "natural_or_challenge",
              "challenge_stratum", "text_sha256", "normalised_text_sha256", "duplicate_group_id"]
    df_s = df.sort_values("external_id").reset_index(drop=True)
    dfp_s = dfp.sort_values("external_id").reset_index(drop=True)
    for col in shared:
        assert col in df.columns and col in dfp.columns, f"Missing {col}"
        assert (df_s[col].astype(str).tolist() == dfp_s[col].astype(str).tolist()), f"Mismatch {col}"


def test_e1_reproducible():
    # Re-run E1 sampling from same candidate frame and assert same IDs
    from v4.external.dedup import compute_dedup, dedup_against_development
    from v4.external.sampling import sample_E1, assign_role_family
    from v4.tests._paths import CORPUS_PATH
    import re
    cand = pd.read_csv(REPO_ROOT / "v4" / "external" / "private" / "reed_candidates_private.csv")
    df = cand.copy()
    df["external_id"] = "reed_" + df["source_posting_id"].astype(str)
    df["role_family"] = df["job_title"].apply(assign_role_family)
    df["role_family_sampling_stratum"] = df["role_family"]
    if "source" not in df.columns:
        df["source"] = "reed"
    df["text_sha256"] = df["job_summary"].apply(lambda x: hashlib.sha256(str(x).encode()).hexdigest())
    df["normalised_text_sha256"] = df["job_summary"].apply(
        lambda x: hashlib.sha256(re.sub(r"\s+", " ", re.sub(r"[^\w ]", "", str(x).lower())).strip().encode()).hexdigest())
    dedup_pool = compute_dedup(df["job_summary"].tolist(), posting_ids=df["external_id"].tolist())
    df["duplicate_group_id"] = dedup_pool["group_ids"]
    corpus = pd.read_csv(CORPUS_PATH)
    dev_texts = corpus["job_summary"].fillna("").astype(str).tolist()
    dev_ids = corpus["posting_id"].astype(str).tolist()
    dedup_dev = dedup_against_development(df["job_summary"].tolist(), df["external_id"].tolist(), dev_texts, dev_ids)
    # remove overlaps (0)
    df2 = df.sort_values("external_id").drop_duplicates(subset=["text_sha256"], keep="first").reset_index(drop=True)
    assert len(df2) == 531, f"Frame size {len(df2)} != 531"
    e1_selected, _ = sample_E1(df2, n_target=200, seed=42, min_per_family=2)
    pub = pd.read_csv(LOCKED, dtype=str)
    locked_e1 = set(pub[pub["natural_or_challenge"] == "natural"]["external_id"])
    repro = set(e1_selected["external_id"])
    assert repro == locked_e1, f"E1 not reproducible diff {sorted(locked_e1 - repro)[:3]}"


def test_e1_has_no_model_dependency():
    txt = (REPO_ROOT / "v4" / "external" / "sampling.py").read_text()
    e1_section = txt.split("def sample_E1")[1].split("def sample_E2")[0]
    assert "A1_pred" not in e1_section
    assert "S1_scores" not in e1_section
    assert "S2_scores" not in e1_section
    assert "S3_scores" not in e1_section
    assert "forbidden" in e1_section.lower() or "model" in e1_section.lower()


def test_e2_counts():
    df = pd.read_csv(LOCKED, dtype=str)
    e2 = df[df["natural_or_challenge"] == "challenge"]
    assert len(e2) == 100
    counts = e2["challenge_stratum"].value_counts().to_dict()
    assert counts.get("C1_lexical_low_coverage_semantic_disagreement", 0) == 40
    assert counts.get("C2_lexical_ambiguity_homonym", 0) == 30
    assert counts.get("C3_role_terminology_edge", 0) == 30
    # disjoint
    e1_ids = set(df[df["natural_or_challenge"] == "natural"]["external_id"])
    e2_ids = set(e2["external_id"])
    assert len(e1_ids & e2_ids) == 0


def test_overlap_counts():
    data = json.loads(OVERLAP_JSON.read_text())
    assert len(data["ids"]) == 100
    assert data["n_E1_overlap"] == 50
    assert data["n_E2_overlap"] == 50
    assert len(set(data["ids"])) == 100
    pub = pd.read_csv(LOCKED, dtype=str)
    locked_ids = set(pub["external_id"])
    assert all(i in locked_ids for i in data["ids"])
    e1 = set(pub[pub["natural_or_challenge"] == "natural"]["external_id"])
    e2 = set(pub[pub["natural_or_challenge"] == "challenge"]["external_id"])
    overlap = set(data["ids"])
    assert len(overlap & e1) == 50
    assert len(overlap & e2) == 50


def test_annotator_a_exact_ids():
    if not WORKBOOK_A.exists():
        pytest.skip("Workbook A private gitignored; not present in ci")
    wb = openpyxl.load_workbook(WORKBOOK_A, data_only=True, read_only=True)
    ws = wb["Annotation"]
    headers = [c.value for c in ws[1]]
    id_col = headers.index("external_id") + 1
    ids = [str(ws.cell(row=r, column=id_col).value).strip() for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=id_col).value]
    pub = pd.read_csv(LOCKED, dtype=str)
    locked = set(pub["external_id"])
    assert len(ids) == 300
    assert set(ids) == locked
    assert len(set(ids)) == 300


def test_annotator_b_exact_overlap_ids():
    if not WORKBOOK_B.exists():
        pytest.skip("Workbook B private")
    wb = openpyxl.load_workbook(WORKBOOK_B, data_only=True, read_only=True)
    ws = wb["Annotation"]
    headers = [c.value for c in ws[1]]
    id_col = headers.index("external_id") + 1
    ids = [str(ws.cell(row=r, column=id_col).value).strip() for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=id_col).value]
    data = json.loads(OVERLAP_JSON.read_text())
    overlap = set(str(x) for x in data["ids"])
    assert len(ids) == 100
    assert set(ids) == overlap


def test_label_cells_blank():
    # Blank check for both workbooks
    for p in [WORKBOOK_A, WORKBOOK_B]:
        if not p.exists():
            continue
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        ws = wb["Annotation"]
        headers = [c.value for c in ws[1]]
        prog_idx = headers.index("programming")
        for r in range(2, ws.max_row + 1):
            for j, cat in enumerate(CATEGORIES):
                v = ws.cell(row=r, column=prog_idx + 1 + j).value
                assert v is None or str(v).strip() == "", f"{p.name} row {r} cat {cat} should be blank, got {v!r}"


def test_no_hidden_model_output():
    for p in [WORKBOOK_A, WORKBOOK_B]:
        if not p.exists():
            continue
        wb = openpyxl.load_workbook(p, data_only=False)
        # No hidden sheets/cols/rows, no forbidden headers
        hidden = [s for s in wb.sheetnames if wb[s].sheet_state == "hidden"]
        assert hidden == [], f"{p.name} has hidden sheets {hidden}"
        ws = wb["Annotation"]
        headers = [c.value for c in ws[1] if c.value]
        forbidden = ["a1", "s1", "s2", "s3", "prediction", "score", "threshold", "confidence"]
        for h in headers:
            assert not any(k in str(h).lower() for k in forbidden), f"{p.name} forbidden header {h}"
        assert sum(1 for d in ws.column_dimensions.values() if d.hidden) == 0
        assert sum(1 for d in ws.row_dimensions.values() if d.hidden) == 0
        # No formulas
        formulas = [c for row in ws.iter_rows() for c in row if c.data_type == "f"]
        assert formulas == [], f"{p.name} has formulas"


def test_taxonomy_order():
    for p in [WORKBOOK_A, WORKBOOK_B]:
        if not p.exists():
            continue
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        ws = wb["Annotation"]
        headers = [c.value for c in ws[1]]
        idx = headers.index("programming")
        seq = headers[idx: idx + 13]
        assert seq == CATEGORIES, f"{p.name} taxonomy mismatch {seq}"


def test_annotation_return_validator():
    # Synthetic valid/invalid tests for validate_annotations
    from v4.external.validate_annotations import validate_workbook
    import tempfile
    import shutil

    # Create synthetic valid workbook with correct IDs but blank labels (should be invalid due to blank? but our validator expects 0/1 -> will report blanks as errors)
    # Instead we test that validator rejects wrong IDs and accepts synthetic filled 0/1 after we fill
    # Prepare a minimal valid file by copying real workbook and filling 0s
    if not WORKBOOK_A.exists():
        pytest.skip("No workbook to base synthetic")

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        # Load real A and fill skill cells with 0, save to temp
        wb = openpyxl.load_workbook(WORKBOOK_A)
        ws = wb["Annotation"]
        headers = [c.value for c in ws[1]]
        prog_idx = headers.index("programming")
        for r in range(2, ws.max_row + 1):
            for j in range(13):
                ws.cell(row=r, column=prog_idx + 1 + j).value = 0
        # also other
        if "other" in headers:
            other_idx = headers.index("other") + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=other_idx).value = 0
        valid_path = td / "valid.xlsx"
        wb.save(valid_path)

        # Should pass when expected IDs are the same 300
        from v4.external.validate_annotations import _load_locked_ids
        locked_ids, _ = _load_locked_ids()
        is_valid, errs = validate_workbook(valid_path, locked_ids, "TEST_VALID")
        assert is_valid, f"Synthetic valid should pass but got {errs}"

        # Invalid: change an ID
        wb2 = openpyxl.load_workbook(valid_path)
        ws2 = wb2["Annotation"]
        ws2.cell(row=2, column=headers.index("external_id") + 1).value = "reed_99999999"
        invalid_path = td / "invalid.xlsx"
        wb2.save(invalid_path)
        is_valid2, errs2 = validate_workbook(invalid_path, locked_ids, "TEST_INVALID")
        assert not is_valid2, "Should reject wrong ID"
        assert any("missing" in e.lower() or "unexpected" in e.lower() for e in errs2)


def test_no_external_evaluation():
    lock = json.loads(LOCK_JSON.read_text())
    assert lock["MODELS_EVALUATED"] is False
    assert lock["LABELS_CREATED"] is False
    assert lock["LABELS_LOCKED"] is False
    # Also ensure no code in v4/external evaluates external gold at this stage (allow agreement)
    for py in (REPO_ROOT / "v4" / "external").rglob("*.py"):
        if py.name in ["agreement.py", "validate_annotations.py"]:
            continue
        content = py.read_text()
        if "EXTERNAL_LABEL_LOCK" in content and "MODELS_EVALUATED" not in content:
            # This would be evaluation
            assert False, f"{py} suggests evaluation on external gold"
